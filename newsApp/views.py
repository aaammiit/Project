from django.shortcuts import render,redirect
from .models import *
import requests
from bs4 import BeautifulSoup
import tenacity
from googlenewsdecoder import new_decoderv1
import datetime
import openpyxl
import smtplib
from django.contrib import messages
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
import os  
from newsTracker import settings
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import time
import re


'''
smtp for sending news file and news on gmail 
'''
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 587
SMTP_USERNAME = 'amistreetecom0101@gmail.com'
SMTP_PASSWORD = 'xmxj ztoo dfvw bxtm'
sender_email = 'amistreetecom0101@gmail.com'
# Receiver_email='newstracker@krimasolutions.com'
Receiver_email='amitdwivedi06503@gmail.com'
subject = "News Alert"


'''
indexPage for a websites landing page 
'''
def indexPage(request):
    return render(request,'index.html')


'''
newshomepage is group of news Categroy and their subgroups 
'''
def newshomePage(request):
    try:
        agency=[]
        categroy=Category.objects.all()
        newsModelData=newsModel.objects.all()
        for i in newsModelData:
            # print(i.newsUrl)
            
            if i not in agency:
                agency.append(i.newsAgency)
        agency1=set(agency)
        data={'data':newsModelData,'categroy':categroy,'agency':agency1,'filterData':categroy}
        return render(request,'newshomepage.html',data)
    except:
        return redirect('/')

def categroyNewsPage(request,id):
    try:
        categroy_id=id
        agency=[]
        categroy = Category.objects.all()    
        # Filter keywords based on category_id
        keyWord = keyWords.objects.filter(category_id=id)
        
        # Retrieve news articles based on filtered keywords
        news_data = newsModel.objects.filter(keyWord__in=keyWord)
        for i in news_data:
            if i not in agency:
                agency.append(i.newsAgency)
        agency1=set(agency)
        data={'keyword':keyWord,'categroy':categroy,'data':news_data,'agency':agency1,'catId':categroy_id}
        return render(request,'newshomepage.html',data)
    except:
        return redirect('/')
    
def keyWordAllNews(request,id):
    try:
        categroy_id=id
        agency=[]
        categroy = Category.objects.all()    
        # Filter keywords based on category_id
        keyWord = keyWords.objects.filter(category_id=id)
        
        # Retrieve news articles based on filtered keywords
        news_data = newsModel.objects.filter(keyWord__in=keyWord)
        for i in news_data:
            if i not in agency:
                agency.append(i.newsAgency)
        agency1=set(agency)
        data={'keyword':keyWord,'categroy':categroy,'data':news_data,'agency':agency1,'catId':categroy_id}
        return render(request,'newshomepage.html',data)
    except:
        return redirect('/')


'''
searchKeywordNews for a search news on google news on basis of keywords , when i give keyword and
time period then this function start scraping first make url then request and response library go website and 
beautigul soap scrap news title , date , what time ago news publish and encoded url decode with use
of googlenewsdecoder and save in a dictonary under a list and redirect in template where news show 
'''

@tenacity.retry(wait=tenacity.wait_exponential(min=4, max=10), stop=tenacity.stop_after_attempt(3))
def searchKeywordNews(request):
    Record=[]
    News=[]
    keyurl=''
    if request.method=='POST':
        did=request.POST.get('catId')
        # print(did)
        url=request.POST.get('url')
        print(url)
        time_period=request.POST.get('time_period')
        keyurl=url
        time_period=time_period
        if url != 'none' and time_period !='none':
            base_url=f'https://news.google.com/search?q={url}20when{time_period}&hl=en-IN&gl=IN&ceid=IN%3Aen'
            # print(base_url)
            response = requests.get(base_url)
            soup = BeautifulSoup(response.content, 'html.parser')
            new_records=[]
            for article in soup.find_all('article'):
                title_tag = article.find_all('a')[1]
                if title_tag:
                    title = title_tag.get_text(strip=True)
                    relative_url = title_tag.get('href')
                    if relative_url and relative_url.startswith('./'):
                        url = 'https://news.google.com' + relative_url[1:]
                        urls=url
                    else:
                        urls=relative_url  
                    try:
                        interval_time = 0
                        decoded_url = new_decoderv1(urls, interval=interval_time)
                        if decoded_url.get("status"):
                            acctualUrl=decoded_url["decoded_url"]
                            listurl=str(acctualUrl).split('/')
                            agency=listurl[2]
                        else:
                            print("Error:", decoded_url["message"])
                    except Exception as e:
                            print(f"Error occurred: {e}")       
                    date_tag = article.find('time')
                    
                    if date_tag:
                        date = date_tag.get('datetime')
                        date = datetime.datetime.strptime(date, '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d')
                        date=datetime.datetime.strptime(date, '%Y-%m-%d').date()
                        date_str = date.isoformat()
                        time_ago=date_tag.text       
                record={
                            'tilte':title,
                            'link':urls,
                            'date':date_str,
                            'ago':time_ago,
                            'agency':agency,
                            'acctualUrl':acctualUrl
                            
                        } 
                if record not in Record:
                    new_records.append(record)
        News.extend(new_records)
    else:
        print('url and time period select None')
    new_list=[]
    sorted_list = sorted(News, key=lambda x: x['date'], reverse=True)
    News1 = sorted_list
    request.session['data']=News1
    if len(News1) !=0:
       data={'data':News1,'url':keyurl,'catId':did}
       return render(request,'fetchNews.html',data)
    else:
        msg={'msg':'Sorry, not any news to fetch.'}
        return render(request,'fetchNews.html',msg)
    
'''
after scraping i want to select specific news and save in our database
'''  
def saveNewsInBucket(request):
    if request.method == 'POST':
        urlkeyword=request.POST.get('url')
        category=request.POST.get('catId')
        categoryId=int(category)
        print(categoryId)
        try:
           newsKeywor=keyWords.objects.get(urlKeyWord=urlkeyword)
        except:
            pass
        try:
            news_ids = request.POST.getlist('news_ids')
            # print(news_ids,type(news_ids))
            data = request.session.get('data')
            for id in news_ids:
                specific_data = data[int(id)]
                model_instance = newsModel(keyWord=newsKeywor,newsTitle=specific_data['tilte'], newsDate=specific_data['date'],newsUrl=specific_data['acctualUrl'],newsAgency=specific_data['agency'])
                model_instance.save()
                print('saved')
            return redirect(f'/categroyNews/{categoryId}')
        except Exception as e:
            print(e)
            return redirect('/')
    else:
        return redirect('/')
        print('err')

def sendFileInEmail(request):
    try:
        data = newsModel.objects.all()
        wb = openpyxl.Workbook()
        ws = wb.active

        # Setting headers for the Excel file
        ws['A1'] = 'Date'
        ws['B1'] = 'Title'
        ws['C1'] = 'Link'
        ws['D1'] = 'keyWord'
        ws['E1'] = 'newsAgency'
        # Populating the Excel file
        for i, record in enumerate(data):  
            ws[f'A{i+2}'] = record.newsDate
            ws[f'B{i+2}'] = record.newsTitle 
            ws[f'C{i+2}'] = record.newsUrl
            ws[f'D{i+2}'] = record.keyWord.keyword
            ws[f'E{i+2}'] = record.newsAgency
        # Filename with current date and time
        date_str = datetime.datetime.now().strftime("%B%d")
        current_time_str = datetime.datetime.now().strftime("%H%M%S")
        filename = f'{date_str}_News_records_{current_time_str}.xlsx'
        wb.save(filename)
        # Preparing email details
        subject = "News Alerts"
        body = "Here are the top and latest news titles and links:\n"
        for record in data:
            body += f"Title: {record.newsTitle},\nPublished Date:{record.newsDate},\n Link: {record.newsUrl}\n newsAgency: {record.newsAgency}\n"

        msg = MIMEMultipart()
        msg['Subject'] = subject
        msg['From'] = sender_email
        msg['To'] = Receiver_email
        msg['Date'] = formatdate(localtime=True)
        try:
            # Attach the text body
            msg.attach(MIMEText(body, 'plain'))
            # Attach the Excel file
            with open(filename, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename={filename}'
                )
                msg.attach(part)
            # Send the email
            server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
            server.starttls()
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(sender_email, Receiver_email, msg.as_string())
            server.quit()
            print('Email sent successfully')
            messages.success(request,'File sent successfully ')
        except Exception as e:
            print(f'Error occurred: {e}')
        print(f'Total records sent: {len(data)}')
        return redirect('/homePage')
    except:
        return redirect('/')
    
def keyWordNews(request, id):
    # Retrieve the keyword object based on the provided ID
    keyWorddata = keyWords.objects.get(id=id)
    catId=keyWorddata.category.id
    keyWorddata1 = keyWords.objects.filter(id=id)  
    # Fetch all keywords belonging to the same category as the selected keyword
    keyWord = keyWords.objects.filter(category_id=keyWorddata.category_id) 
    # Fetch all news articles that are linked to any of the keywords in the same category
    news_data = newsModel.objects.filter(keyWord__in=keyWorddata1)  # Updated field name
    agency = {i.newsAgency for i in news_data}
    # Prepare data for the template
    data = {
        'keyword': keyWord,
        'data': news_data,
        'agency': agency , # This is already unique because it's a set
        'catId':catId
    }
    return render(request, 'newshomepage.html', data)


chrome_path = f"{settings.BASE_DIR}/chrome-win64/chrome-win64/chrome.exe"
chromedriver_path = f"{settings.BASE_DIR}/chrome-win64/chromedriver-win64/chromedriver.exe"

def newsUrlScrap(request,id):
    newsArticles = newsArticle()
    chrome_options = Options()
    chrome_options.binary_location = chrome_path
    service = Service(chromedriver_path)
    driver = None
    try:
        driver = webdriver.Chrome(service=service, options=chrome_options)
        data = newsModel.objects.filter(id=id).first()
        if not data:
            return redirect('/')
        url = data.newsUrl
        driver.get(url)
        time.sleep(2) 
        highlight_js = """
        window.clickedTexts = [];
        window.clickedTables = [];
        window.clickedElements = new Set();

        document.querySelectorAll('*').forEach(el => {
            el.addEventListener('mouseover', function(event) {
                if (!window.clickedElements.has(el)) {
                    event.target.style.outline = '2px solid red';
                }
            });

            el.addEventListener('mouseout', function(event) {
                if (!window.clickedElements.has(el)) {
                    event.target.style.outline = '';
                }
            });

            el.addEventListener('click', function(event) {
                event.preventDefault();
                const text = event.target.innerText.trim();

                if (el.tagName.toLowerCase() === 'table') {
                    // If a table is clicked, extract its structure
                    let tableData = [];
                    el.querySelectorAll('tr').forEach(row => {
                        let rowData = [];
                        row.querySelectorAll('th, td').forEach(cell => {
                            rowData.push(cell.innerText.trim());
                        });
                        tableData.push(rowData);
                    });
                    
                    window.clickedTables.push({ structure: tableData });
                    window.clickedElements.add(el);
                    event.target.style.outline = '2px solid blue';
                    console.log("Added table data:", tableData);
                } else if (text) {
                    if (!window.clickedElements.has(el)) {
                        window.clickedTexts.push(text);
                        window.clickedElements.add(el);
                        event.target.style.outline = '2px solid blue';
                        console.log("Added text:", text);
                    } else {
                        window.clickedTexts = window.clickedTexts.filter(t => t !== text);
                        window.clickedElements.delete(el);
                        event.target.style.outline = '';
                        console.log("Removed text:", text);
                    }
                }
            });
        });

        var doneButton = document.createElement('button');
        doneButton.innerText = "Done";
        doneButton.id = "doneButton";
        doneButton.style.position = "fixed";
        doneButton.style.top = "10px";
        doneButton.style.right = "10px";
        doneButton.style.zIndex = 10000;
        doneButton.style.backgroundColor = "green";
        doneButton.style.color = "white";
        doneButton.style.padding = "10px";
        doneButton.style.border = "none";
        doneButton.style.cursor = "pointer";
        doneButton.onclick = function() {
            window.clickedTexts.push("done");
        };
        document.body.appendChild(doneButton);
        """
        driver.execute_script(highlight_js)
        stored_texts = []
        print("Hover over elements and click to store their text or table data. Click again to remove. Click 'Done' to finish.")
        while True:
            time.sleep(1)  # Wait for user interaction
            clicked_texts = driver.execute_script("return window.clickedTexts;")
            if clicked_texts and "done" in clicked_texts:
                print("Exiting... 'Done' button clicked.")
                break
            stored_texts = clicked_texts
            final_text = list(set(stored_texts))
            cleaned_text = ' '.join(final_text)
            cleaned_text = re.sub(r'[\n\t\s]+', ' ', cleaned_text).strip()
        if cleaned_text:
            newsArticles.newsarticle = cleaned_text
            newsArticles.news = data    
            newsArticles.save()
            data.isScraped = True
            data.save()
            print('Saved successfully.')
            return redirect('/homePage')
        else:
            print("No text found to save.")
            return redirect('/homePage')
    except Exception as e:
        print(f"An error occurred: {e}")
        return redirect('/')
    finally:
        if driver:
            driver.quit()  # Ensure the driver quits to free up resources

def newsArticles(request):
    article=newsArticle.objects.all()
    data={'article':article}
    return render(request,'newsArticles.html',data)

def deleteAllData(request):
    data=newsModel.objects.all()
    try:
        data.delete()
        return redirect('/homePage')
    except newsModel.DoesNotExist as e:
        print(e)
        return redirect('/homePage')
    
def removeNews(request,id):
    news=newsModel.objects.get(id=id)
    news.delete()
    return redirect('/homePage')

def editArticle(request,id):
    try:
        article=newsArticle.objects.get(id=id)
        if request.method=='POST':
            newsarticle=request.POST.get('articleContent')
            newsTitle=request.POST.get('title')
            newsaDate=request.POST.get('date')
            article.newsarticle=newsarticle
            article.news.newsTitle=newsTitle
            article.news.newsDate=newsaDate
            article.isEdited=True
            article.save()
            return redirect('/newsArticles')
        else:
            data={'article':article}
            return render(request,'newsArticlesEdit.html',data)
    except:
        return redirect('/')


def gptSummary(request,id):
    try:
        article=newsArticle.objects.get(id=id)
        data={'article':article}
        return render(request,'newsArticlesSummary.html',data)
    except:
        return redirect('/newsArticles')
    

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

import json

@csrf_exempt
def save_content(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)  # Parse the incoming JSON data
            article_id = data.get('articleId')  # Get the article ID
            
            content = data.get('content')  # Get the content from the Quill editor

            # Fetch the article object and update its fields
            article = newsArticle.objects.get(id=article_id)
            article.newsarticle = content
            article.save()  # Save the updated article

            return JsonResponse({'status': 'success', 'message': 'Content saved successfully!'})
        except newsArticle.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Article not found'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from .models import newsArticle  # Assuming the newsArticle model is imported from your models

@csrf_exempt
def update_article_content(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)  # Parse the incoming JSON data
            article_id = data.get('articleId')  # Get the article ID
            
            content = data.get('content')  # Get the content from the Quill editor

            # Fetch the article object and update its fields
            article = newsArticle.objects.get(id=article_id)
            article.gptSummary = content
            article.save()  # Save the updated article

            return JsonResponse({'status': 'success', 'message': 'Content updated successfully!'})
        except newsArticle.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Article not found'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)




